#! /usr/bin/python
# -*- coding: utf-8 -*-

import xlrd
from openpyxl import Workbook
import csv
import json
import urllib2
from utils import timer
import utils

geocode_url = 'http://api.map.baidu.com/geocoder/v2/'
placeAPI_url = 'http://api.map.baidu.com/place/v2/search'
ak_mcm = 'Rl3cuYvuDMBES9TULIixBWAIOi0ES7BN'
output_col_name = {1:"编号", 2:"订单号", 3:"经度", 4:"纬度", 5:"地址"}


class ExcueteFile(object):

    def __init__(self,input_file, output_file, anchor_col_letter, address_cel_letter, sheet_num, start_row, chinese_code, method="geocoder"):
        self.input_file = input_file
        self.output_file = output_file
        self.anchor_col = utils.get_excel_col_number(anchor_col_letter)
        self.address_col = utils.get_excel_col_number(address_cel_letter)
        self.start_row = start_row
        self.sheet_num = sheet_num
        self.chinese_code = chinese_code
        self.method = method

    @staticmethod
    def get_url_for_geocode(address):
        option = "city=广州&output=json&ret_coordtype=gcj02ll&address="  # &coordtype=gcj02ll
        print 'address: ', address
        return geocode_url + "?" + option + address + "&ak=" + ak_mcm

    @staticmethod
    def get_url_for_placeAPI(address):
        print 'address: ', address
        option = "region=广州&scope=1&ret_coordtype=gcj02ll&output=json&query="  # ret_coordtype=gcj02ll&
        return placeAPI_url + "?" + option + address + "&ak=" + ak_mcm

    def isStreetAddress(self, str):
        if self.chinese_code == "GBK":
            if str.count("区".decode("utf-8").encode("GBK")) > 0 or str.count("市".decode("utf-8").encode("GBK")) > 1:
                return True
        elif self.chinese_code == "unicode":
            if str.count("区".decode("utf-8")) > 0 or str.count("市".decode("utf-8")) > 1:
                return True
        else:
            raise ("unknow chinese code")
        return False

    def isRedundancyStreetAddress(self, str):
        if self.chinese_code == "GBK":
            if str.count("区".decode("utf-8").encode("GBK")) > 1 or str.count("市".decode("utf-8").encode("GBK")) > 2:
                return str[22:]
        elif self.chinese_code == "unicode":
            if str.count("区".decode("utf-8")) > 1 or str.count("市".decode("utf-8")) > 2:
                return str[6:]
        else:
            raise ("unknow chinese code")
        return str

    def strWork(self, str):
        str = str.replace(" ", "").replace('"', '').replace("=", "").replace('\r\n',"").replace("-", "")  # 去除空格和
        str = self.isRedundancyStreetAddress(str)
        return str

    def getAxisUseGeocode(self, address):
        if self.chinese_code == "GBK":
            url = self.get_url_for_geocode(address.decode("GBK").encode("utf-8"))
        elif self.chinese_code == "unicode":
            url = self.get_url_for_geocode(address.encode("utf-8"))
        else:
            raise ("unknow chinese code")
        print url

        temp = urllib2.urlopen(url, timeout=5)
        str = temp.read()
        data = json.loads(str)

        lng = data["result"]["location"]["lng"]
        lat = data["result"]["location"]["lat"]

        print lng, lat
        return [lng, lat]

    def getAxisUsePlaceAPI(self, address):
        if self.chinese_code == "GBK":
            url = self.get_url_for_placeAPI(address.decode("GBK").encode("utf-8"))
        elif self.chinese_code == "unicode":
            url = self.get_url_for_placeAPI(address.encode("utf-8"))
        else:
            raise ("unknow chinese code")
        print url

        temp = urllib2.urlopen(url, timeout=5)
        str = temp.read()
        data = json.loads(str)

        lng = data["results"][0]["location"]["lng"]
        lat = data["results"][0]["location"]["lat"]

        print lng, lat

        return [lng, lat]

    @timer
    def getAxis(self, address):
        address = self.strWork(address)

        if self.isStreetAddress(address):
            return self.getAxisUseGeocode(address)
        else:
            return self.getAxisUsePlaceAPI(address)

    def read_csv(self, filepath, start_row):
        lst = []

        csvReader = csv.reader(file(filepath, 'rb'))
        for i in xrange(start_row-1):
            csvReader.next()
        for row in csvReader:
            lst.append([row[self.anchor_col], row[self.address_col]])
        return lst

    def read_excel(self, filepath, start_row, sheet_num):
        lst = []
        with xlrd.open_workbook(filepath) as workbook:
            #sheet = workbook.sheet_by_index(sheet_num)
            sheet = workbook.sheet_by_name("营业厅负责人".decode("utf-8"))
            rown = sheet.nrows

            print self.anchor_col, self.address_col

            for r in range(start_row-1, rown):
                # for i in range(2):
                #     print sheet.cell_value(r, i)
                # print " "
                lst.append([sheet.cell_value(r, self.anchor_col),
                            sheet.cell_value(r, self.address_col)])
        return lst

    def read(self, filepath, start_row):
        if filepath.endswith(".csv"):
            return self.read_csv(filepath, start_row)
        elif filepath.endswith((".xls", ".xlsx")):
            return self.read_excel(filepath, start_row, self.sheet_num)

    def multiSearch(self):
        addresslist = self.read(self.input_file, self.start_row)

        axislist = []
        i = self.start_row
        print "filelength= ", len(addresslist)
        for address in addresslist:

            try:

                axis = self.getAxis(address[1])
                print i, axis
                axislist.append([i, address[0], axis[0], axis[1], address[1]])
            except:
                print i, address
            finally:
                i += 1

        return axislist

    @staticmethod
    def writeExcel(axislist, filepath):
        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value=output_col_name[1])
        ws.cell(row=1, column=2, value=output_col_name[2])
        ws.cell(row=1, column=3, value=output_col_name[3])
        ws.cell(row=1, column=4, value=output_col_name[4])
        ws.cell(row=1, column=5, value=output_col_name[5])

        r = 1
        for axis in axislist:
            print r
            r += 1
            ws.cell(row=r, column=1, value=axis[0])
            ws.cell(row=r, column=2, value=axis[1])
            ws.cell(row=r, column=3, value=axis[2])
            ws.cell(row=r, column=4, value=axis[3])
            ws.cell(row=r, column=5, value=axis[4])

        wb.save(filepath)

    def run(self):
        axislist = self.multiSearch()
        self.writeExcel(axislist, self.output_file)


if __name__ == "__main__":
    ef = ExcueteFile(   #输入文件路径
                        r"F:\视频组\投诉\营业线通讯录（20170405）.xlsx".decode("utf-8").encode("GBK"),
                        #输出文件路径
                        r"F:\视频组\投诉\营业线通讯录（20170405）_result.xlsx".decode("utf-8").encode("GBK"),
                        "A",         #锚定列
                        "O",         #地址列
                        1,           #sheet号
                        2,           #数据起始行号
                        "unicode"   #文件中文编码，一般是GBK，或者unicode
                        )
    #ef.getAxis("广州市海珠区沥窖村沥窖大街76号之6-7铺".decode("utf-8"))
    ef.run()
    # ExcueteFile(r"F:\视频组\地址经纬度\bilibili广州.csv".decode("utf-8").encode("GBK"),
    #         r"F:\视频组\地址经纬度\bilibili广州_result.xlsx".decode("utf-8").encode("GBK"),
    #         "A",
    #         "Q",
    #         2,
    #         "GBK").run()
    # http://api.map.baidu.com/geocoder/v2/?city=广州&output=json&ret_coordtype=gcj02ll&address=广东省经济经贸职业技术学校&ak=Rl3cuYvuDMBES9TULIixBWAIOi0ES7BN
    # http://api.map.baidu.com/place/v2/search?query=广东省经济经贸职业技术学校&region=广州&scope=1&output=json&ret_coordtype=gcj02ll&ak=Rl3cuYvuDMBES9TULIixBWAIOi0ES7BN