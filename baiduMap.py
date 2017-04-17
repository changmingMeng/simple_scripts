#! /usr/bin/python
# -*- coding: utf-8 -*-

import xlrd
from openpyxl import Workbook
import csv
import json
import urllib2
from utils import timer
import utils
import chardet

url = 'http://api.map.baidu.com/geocoder/v2/'
placeAPIUrl = 'http://api.map.baidu.com/place/v2/search'
ak_lst = [
          'Rl3cuYvuDMBES9TULIixBWAIOi0ES7BN',
          '38ADFrK5VVdptqmai3HnjN78p4RjjQeB',
          'UYK8YTMm8ly6Sz1aMcnXsXTO9ohni2ex',
          'YjkhAC8a6XEZqtCN0kSGL8GI1OVLTCly',
          'Hej6HgVATRb3c1brHnoLh9nnfmYiP6YX',
          ]
csv_order_col_number = utils.get_excel_col_number('A')
csv_address_col_number = utils.get_excel_col_number('Q')
excel_order_col_number = utils.get_excel_col_number('P')
excel_address_col_number = utils.get_excel_col_number('G')
#output_col_name = {1:"订单号", 2:"经度", 3:"纬度"}
output_col_name = {1:"编号", 2:"订单号", 3:"经度", 4:"纬度"}
chinese_code = "unicode"
class baiduMap(object):

    @staticmethod
    def getUrl(address, ak=ak_lst[0]):
        option = "city=广州&output=json&ret_coordtype=gcj02ll&address="#&coordtype=gcj02ll
        print 'address: ', address
        return url+"?"+option+address+"&ak="+ak

    @staticmethod
    def getUrlForPlaceAPI(address, ak=ak_lst[0]):
        print 'address: ', address.decode("GBK").encode("utf-8")
        option = "region=广州&scope=1&ret_coordtype=gcj02ll&output=json&query="#ret_coordtype=gcj02ll&
        return placeAPIUrl+"?"+option+address+"&ak="+ak

    @staticmethod
    def strWork(str):
        str = str.replace(" ","").replace("-","").replace('"', '').replace("=", "")#去除空格和-
        if chinese_code == "GBK":
            if str.count("区".decode("utf-8").encode("GBK"))>1 or str.count("市".decode("utf-8").encode("GBK"))>2:
                str = str[22:]
        elif chinese_code == "unicode":
            if str.count("区".decode("utf-8"))>1 or str.count("市".decode("utf-8"))>2:
                str = str[22:]
        else:
            raise ("unknow chinese code")
        return str

    @classmethod
    @timer
    def getAxis(cls, address, ak=ak_lst[0], mode="geocoder"):
        #print 'address: ', address.decode('GBK').encode('utf-8')
        address = cls.strWork(address)
        #print 'new address: ', address.decode('GBK').encode('utf-8')
        #return [0, 0]
        if mode == "geocoder":
            if chinese_code == "GBK":
                url = cls.getUrl(address.decode("GBK").encode("utf-8"), ak)
            elif chinese_code == "unicode":
                url = cls.getUrl(address.encode("utf-8"), ak)
            else:
                raise ("unknow chinese code")
        #url = cls.getUrlForPlaceAPI(address, ak)
            print url
            temp = urllib2.urlopen(url, timeout=2)
            str = temp.read()
            data = json.loads(str)

            lng = data["result"]["location"]["lng"]
            lat = data["result"]["location"]["lat"]
        else:
            url = cls.getUrlForPlaceAPI(address.decode("GBK").encode("utf-8"), ak)
            print url
            temp = urllib2.urlopen(url, timeout=2)
            str = temp.read()
            data = json.loads(str)
            lng = data["results"][0]["location"]["lng"]
            lat = data["results"][0]["location"]["lat"]

        return [lng, lat]


def readCSV(filepath):
    lst = []

    csvReader = csv.reader(file(filepath, 'rb'))
    for i in xrange(3):
        csvReader.next()
    for row in csvReader:
        lst.append([row[csv_order_col_number], row[csv_address_col_number]])
    return lst

def readExcel(filepath):
    lst = []
    with xlrd.open_workbook(filepath) as workbook:
        sheet = workbook.sheet_by_index(0)
        rown = sheet.nrows

        for r in range(rown):
            lst.append([sheet.cell_value(r, excel_order_col_number),
                        sheet.cell_value(r, excel_address_col_number)])
    return lst

def getAddressList(filepath):
    if filepath.endswith(".csv"):
        return readCSV(filepath)
    elif filepath.endswith((".xls", ".xlsx")):
        return readExcel(filepath)


def multiSearch(filepath):
    addresslist = getAddressList(filepath)
    #print addresslist
    #print len(addresslist)
    axislist = []
    i = 1
    print "filelength= ", len(addresslist)
    for address in addresslist:
        try:
            axis = baiduMap.getAxis(address[1], ak_lst[0], "geocoder")
            print i, axis
            axislist.append([i,address[0],axis[0],axis[1]])
        except:
            print i, address
            #axislist.append([i,address[0], address[1], "未取到"])
        finally:
            i += 1
            #if

    return axislist

def writeExcel(axislist, filepath):
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value=output_col_name[1])
    ws.cell(row=1, column=2, value=output_col_name[2])
    ws.cell(row=1, column=3, value=output_col_name[3])
    ws.cell(row=1, column=4, value=output_col_name[4])

    r = 1
    for axis in axislist:
        print r
        r += 1
        ws.cell(row=r, column=1, value=axis[0])
        ws.cell(row=r, column=2, value=axis[1])
        ws.cell(row=r, column=3, value=axis[2])
        ws.cell(row=r, column=4, value=axis[3])

    wb.save(filepath)


def test(inputfilepath, outputfilepath):
    writeExcel(multiSearch(inputfilepath), outputfilepath)


if __name__ == "__main__":
    #print multiSearch(r"F:\视频组\投诉\3-30\address.csv".decode("utf-8").encode("GBK"))
    #print baiduMap.getAxis("华南理工大学广州学院",ak_lst[0])
    #print getAddressList(r"F:\视频组\投诉\3-30\address.xlsx".decode("utf-8").encode("GBK"))
    test(r"F:\视频组\投诉\4-13\B2I投诉工单-广州（原始）1.xlsx".decode("utf-8").encode("GBK"),
         r"F:\视频组\投诉\4-13\B2I投诉工单-广州（原始）1_result.xlsx".decode("utf-8").encode("GBK"))
    #test(r"F:\视频组\地址经纬度\bilibili广州.csv".decode("utf-8").encode("GBK"),
    #     r"F:\视频组\地址经纬度\bilibili广州_result.xlsx".decode("utf-8").encode("GBK"))


    #[113.36763702320629, 23.14023770937979]
    #[113.36763702320629, 23.14023770937979]

    #[113.17212192764, 23.435253319248]
    #[113.178584, 23.441531]
    #print getAddressList(r"F:\视频组\投诉\3-30\address.csv".decode("utf-8").encode("GBK"))
    #print "广东广州市黄埔区  广东省广州  市花都区迎宾大道名高思埠大厦".replace(" ","")
    #http://api.map.baidu.com/place/v2/search?query=%E4%B8%8A%E7%A4%BE&scope=1&region=%E5%B9%BF%E5%B7%9E&output=json&ret_coordtype=gcj02ll&ak=UYK8YTMm8ly6Sz1aMcnXsXTO9ohni2ex
    # http://api.map.baidu.com/place/v2/search?query=广东广州市黄埔区广东省广州市花都区迎宾大道名高思埠大厦&scope=1&region=%E5%B9%BF%E5%B7%9E&output=json&ret_coordtype=gcj02ll&ak=UYK8YTMm8ly6Sz1aMcnXsXTO9ohni2ex
    # http://api.map.baidu.com/place/v2/search?query=广东广州市黄埔区广东省广州市花都区迎宾大道名高思埠大厦&region=广州&scope=1&output=json&ret_coordtype=gcj02ll&ak=UYK8YTMm8ly6Sz1aMcnXsXTO9ohni2ex
    # address = "广东广州市黄埔区广东省 广州市 越秀区    一德路498号一楼a5档"
    # print baiduMap.strWork(address)