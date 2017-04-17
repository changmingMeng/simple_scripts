# coding: utf-8

import os
from os.path import join
import psycopg2
from time import time
import xlrd
import datetime
from openpyxl import Workbook

def timer(func):
    def warpper(*args, **kw):
        start = datetime.datetime.now()
        func(*args, **kw)
        end = datetime.datetime.now()
        print end-start
    return warpper

class Utils(object):
    def GetCellId(self, desc, str):
        return desc[desc.find(str)+len(str):]

    def IsStringLike(self, anobj):
        try:
            anobj.lower() + anobj + ''
        except:
            return False
        else:
            return True

    def GetDateAndTime(self, dateAndTime):
        dateTime = [int(dateAndTime[0:4]), int(dateAndTime[5:7]), int(dateAndTime[8:10]),
                    int(dateAndTime[11:13]), int(dateAndTime[14:16])]
        return dateTime

    def GetDateAndTimeForPostgresql(self, dateAndTime):
        timeNum = self.GetDateAndTime(dateAndTime)
        date = psycopg2.Date(timeNum[0], timeNum[1], timeNum[2])
        time = psycopg2.Time(timeNum[3], 0, 0)
        return [date, time]

    def GetDateAndTimeNormal(self, dateAndTime):
        date, time = self.GetDateAndTimeForPostgresql(dateAndTime)
        return [self.pgDateToStr(date), self.pgTimeToStr(time)]

    def pgDateToStr(self, date):
        return str(date)[1:11]

    def pgTimeToStr(self, time):
        return str(time)[1:9]

    def dateTimeIdNTToStr(self, dateTimeId):
        return dateTimeId[0] + dateTimeId[1] + dateTimeId[2] + dateTimeId[3]

    def StrToDateTimeIdNT(self, dateTimeId):
        return [dateTimeId[0:10], dateTimeId[10:18], dateTimeId[18:20], dateTimeId[20:]]

    @staticmethod
    def tuple2Sqlite3Timestring(t):
        """(2017,1,1)型元组转换成Sqlite3的“YYYY-MM-DD”型字符串"""
        year, month, day = t
        if month/10 is 0:
            m = "0"+str(month)
        else:
            m = str(month)

        if day/10 is 0:
            d = "0"+str(day)
        else:
            d = str(day)

        return str(year)+"-"+m+"-"+d

    @staticmethod
    def exceldate_to_postgredate(date):
        """excel读入的日期为一个数字，转化成postgreSQL.Date类型"""
        return psycopg2.Date(*xlrd.xldate_as_tuple(date, 0)[:3])

    @staticmethod
    def strdate_to_postgredate(date):
        """2017-01-01形式的字符串转化成postgreSQL.Date类型"""
        b = tuple(date.split('-'))
        return psycopg2.Date(int(b[0]), int(b[1]), int(b[2]))

    @staticmethod
    def lst_of_lst_distince_by_col(lst_of_lst, col_num):
        result_lst = []
        distinct_lst = []

        for lst in lst_of_lst:
            if lst[col_num] not in distinct_lst:
                distinct_lst.append(lst[col_num])
                result_lst.append(lst)

        return result_lst


def get_sequence_num_of_letter(letter):
    #print letter
    return ord(letter.upper())-ord('A')+1

def get_excel_col_number(str):
    if len(str) == 1:
        return get_sequence_num_of_letter(str[0])-1
    elif len(str) == 2:
        return get_sequence_num_of_letter(str[0])*26 +get_sequence_num_of_letter(str[1])-1
    else:
        raise("too many letters")

def write_lst_of_lst_to_excel(filepath,lst_name, lst_of_lst):
    wb = Workbook()
    ws = wb.active

    for i in range(len(lst_name)):
        ws.cell(row=1, column=i+1, value=lst_name[i])

    for i in range(len(lst_of_lst)):
        for j in range(len(lst_of_lst[i])):
            ws.cell(row=i+2, column=j+1, value=lst_of_lst[i][j])

    wb.save(filepath)


def timer(func):
    def warpper(*args, **kw):
        tic=time()
        result=func(*args, **kw)
        toc=time()
        print "%f seconds has passed"%(toc-tic)
        return result
    return warpper

if __name__ == "__main__":
    # a = "测试RNC/BSC6900UCell:Label=W测试RNC基站1, CellID=9991"
    # a2 = "GZRNC15/BSC6900UCell:Label=W夏茅工业区1, CellID=26331"
    # b = "CellID="
    ut = Utils()
    # print ut.GetCellId(a, b)
    # print ut.GetCellId(a2, b)
    # dateAndTime = "2016-11-27 23:00"
    # print ut.GetDateAndTime(dateAndTime)
    #
    # print psycopg2.Time(16, 12, 14)
    # print str(psycopg2.Time(16, 12, 14))[1:9]
    # print ut.pgTimeToStr(psycopg2.Time(16, 12, 14))
    # list = ['2016-11-28', '22:00:00', '2G','4600125046183']
    # str = ut.dateTimeIdNTToStr(list)

    # print ut.tuple2Sqlite3Timestring((2017, 1, 1))
    # print ut.tuple2Sqlite3Timestring((2017, 11, 1))
    # print ut.tuple2Sqlite3Timestring((2017, 1, 11))
    # print type("七所搬迁-1")
    # print type("七所搬迁-1") is str
    # print type(1.5) is str

    date = datetime.date.today()
    print type(date)
    print str(date).replace("-","")